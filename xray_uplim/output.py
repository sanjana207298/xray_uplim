"""
xray_uplim.output
-----------------
Shared helpers for writing pipeline results to disk.

write_results_xlsx(rows, fieldnames, xlsx_path, text_cols)
    Write a list of row dicts to an Excel workbook (.xlsx).
    Columns listed in text_cols are formatted as plain text so that
    Excel never strips leading zeros from observation IDs.

This is called alongside write_results_csv() in every telescope pipeline.
openpyxl is required; if not installed the function warns and returns None.
"""

import warnings


def write_results_xlsx(rows, fieldnames, xlsx_path, text_cols=('obsid',)):
    """
    Write pipeline result rows to an Excel workbook.

    All columns are written as-is; columns named in text_cols are
    explicitly formatted as text (number_format '@'), which prevents
    Excel from stripping leading zeros from numeric-looking strings
    such as Swift obsids ('03000397004').

    Parameters
    ----------
    rows       : list of dict   — same rows passed to csv.DictWriter
    fieldnames : list of str    — column order
    xlsx_path  : str            — output path (created/overwritten)
    text_cols  : tuple of str   — columns to force-format as text

    Returns
    -------
    xlsx_path if written successfully, None if openpyxl is unavailable.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        warnings.warn(
            "openpyxl is not installed — Excel (.xlsx) output skipped.\n"
            "Install with:  pip install openpyxl",
            UserWarning, stacklevel=2)
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "upper_limits"

    # ── Header row ────────────────────────────────────────────────────────
    header_font  = Font(bold=True)
    header_fill  = PatternFill("solid", fgColor="D9E1F2")   # light blue
    header_align = Alignment(horizontal="center")

    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # ── Data rows ─────────────────────────────────────────────────────────
    text_col_indices = {
        col_idx for col_idx, name in enumerate(fieldnames, start=1)
        if name in text_cols
    }

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(fieldnames, start=1):
            value = row.get(name, '')
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in text_col_indices:
                # Force text format: Excel will not interpret the value
                # as a number, preserving leading zeros in obsids.
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal="left")

    # ── Auto-size columns (capped at 40 chars) ────────────────────────────
    for col_idx, name in enumerate(fieldnames, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(name)),
            *(len(str(row.get(name, ''))) for row in rows),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

    # ── Freeze top row ────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    return xlsx_path
